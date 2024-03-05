from dataclasses import dataclass
import logging
import time
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import units
from ConfigsManager import ConfigsManager
import datetime
from datetime import timedelta
from stravalib.model import Activity
from stravalib.attributes import EntityCollection
import sys
from StravaService import StravaService

from model import DayActivities, WeeksActivities


log = logging.getLogger("APP." + __name__)
class ExcelWriter:
    HEIGHT_KEY = "height"
    WEIGHT_KEY = "weight"
    AGE_KEY = "age"
    SEX_KEY = "sex"
    MAX_HEART_RATE_KEY = "max_heart_rate"
    ZONE_ONE = "zone1"
    ZONE_TWO = "zone2"
    ZONE_THREE = "zone3"
    ZONE_FOUR = "zone4"
    FILE_NAME = "activities.xlsx"
    OTHER_RECOVERY_TIME_COLUMN = 12
    ZONE_1_COLUMN_NUMBER = 13
    ZONE_2_COLUMN_NUMBER = 14
    ZONE_3_COLUMN_NUMBER = 15
    ZONE_4_COLUMN_NUMBER = 16
    ZONE_5_COLUMN_NUMBER = 17
    

    CALORIES_KEY = "calories"

    CALORIES_COLUMN = 10
    REPORT_COLUMN = 11

    __stravaService : StravaService
    __configs = {}
    __heart_rate_zones : list[int] = []
    __should_fully_rewrite = False
    __bmr : float = 0
    __work_book : Workbook

    def __init__(self, stravaService: StravaService):
        self.__work_book = load_workbook(self.FILE_NAME)
        self.__configs = ConfigsManager().read_configs()
        max_heart_rate = self.__configs[self.MAX_HEART_RATE_KEY]
        self.__heart_rate_zones = [max_heart_rate * self.__configs[self.ZONE_ONE], 
                                   max_heart_rate * self.__configs[self.ZONE_TWO], 
                                   max_heart_rate * self.__configs[self.ZONE_THREE], 
                                   max_heart_rate * self.__configs[self.ZONE_FOUR],
                                   max_heart_rate]
        self.__stravaService = stravaService

    def __to_float(self, value) -> float:
        try:
            return float(value)
        except ValueError:
            return 0

    def __to_int(self, value) -> int:
        try:
            return int(value)
        except ValueError:
            return 0
        
    def __fill_sheet_with_default(self, work_sheet: Worksheet):
        #actually you can, but result is not quaranteed ;)
        red_font = Font(color="FF0000")
        work_sheet.cell(2, 2).value = "Don\'t modify this column/row"
        work_sheet.cell(2, 2).alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        work_sheet.cell(2, 2).font = red_font
        work_sheet.cell(2, 3).value = "Monday"
        work_sheet.cell(2, 4).value = "Tuesday"
        work_sheet.cell(2, 5).value = "Wednesday"
        work_sheet.cell(2, 6).value = "Thursday"
        work_sheet.cell(2, 7).value = "Friday"
        work_sheet.cell(2, 8).value = "Saturday"
        work_sheet.cell(2, 9).value = "Sunday"
        work_sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=9)
        work_sheet.cell(1, 3).value = "Every activity has id. Id could be used to access activity details with link https://www.strava.com/activities/{id}"
        
        work_sheet.cell(2, 10).value = "Calories"
        work_sheet.cell(2, 11).value = "Report"
        work_sheet.cell(2, self.OTHER_RECOVERY_TIME_COLUMN).value = "Other recovery activities(Yoga, stretching, etc.). Manual entry in minutes if not a part of activity"
        work_sheet.cell(2, self.ZONE_1_COLUMN_NUMBER).value = f"Zone 1 < {round(self.__heart_rate_zones[0])} bpm"
        work_sheet.cell(2, self.ZONE_2_COLUMN_NUMBER).value = f"Zone 2 {round(self.__heart_rate_zones[0])}-{round(self.__heart_rate_zones[1])} bpm"
        work_sheet.cell(2, self.ZONE_3_COLUMN_NUMBER).value = f"Zone 3 {round(self.__heart_rate_zones[1])}-{round(self.__heart_rate_zones[2])} bpm"
        work_sheet.cell(2, self.ZONE_4_COLUMN_NUMBER).value = f"Zone 4 {round(self.__heart_rate_zones[2])}-{round(self.__heart_rate_zones[3])} bpm"
        work_sheet.cell(2, self.ZONE_5_COLUMN_NUMBER).value = f"Zone 5 {round(self.__heart_rate_zones[3])}-{self.__heart_rate_zones[4]} bpm"
        work_sheet.merge_cells(start_row=1, start_column=self.ZONE_1_COLUMN_NUMBER, end_row=1, end_column=self.ZONE_5_COLUMN_NUMBER)
        work_sheet.cell(1, self.ZONE_1_COLUMN_NUMBER).value = "Zone values are taken from full activity details. If there are several laps estimation is made by laps HR otherwise by splits HR"
        work_sheet.cell(1, self.ZONE_1_COLUMN_NUMBER).alignment = Alignment(wrap_text=True, vertical='top')
        work_sheet.column_dimensions['B'].width = 15
        work_sheet.column_dimensions['C'].width = 25
        work_sheet.column_dimensions['D'].width = 25
        work_sheet.column_dimensions['E'].width = 25
        work_sheet.column_dimensions['F'].width = 25
        work_sheet.column_dimensions['G'].width = 25
        work_sheet.column_dimensions['H'].width = 25
        work_sheet.column_dimensions['I'].width = 25
        work_sheet.column_dimensions['K'].width = 35
        work_sheet.column_dimensions['L'].width = 30
        work_sheet.row_dimensions[1].height = 60
        work_sheet.row_dimensions[2].height = 50

    def __apply_style(self, work_sheet: Worksheet):
        iterator = 2
        while work_sheet.cell(iterator, 2).value is not None:
            work_sheet.cell(iterator, self.ZONE_1_COLUMN_NUMBER).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type = "solid")
            work_sheet.cell(iterator, self.ZONE_2_COLUMN_NUMBER).fill = PatternFill(start_color="6D9EEB", end_color="6D9EEB", fill_type = "solid")
            work_sheet.cell(iterator, self.ZONE_3_COLUMN_NUMBER).fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type = "solid")
            work_sheet.cell(iterator, self.ZONE_4_COLUMN_NUMBER).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
            work_sheet.cell(iterator, self.ZONE_5_COLUMN_NUMBER).fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type = "solid")
            for column_iterator in range(3, 11):
                work_sheet.cell(iterator, column_iterator).alignment = Alignment(wrap_text=True, vertical='top')
            for column_iterator in range(12, 18):
                work_sheet.cell(iterator, column_iterator).alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                
            iterator = iterator + 1

    def __get_worksheet(self, work_book : Workbook) -> Worksheet:
        activityLogSheetName = "Activity Log"
        work_sheet = None 
        if activityLogSheetName not in work_book.get_sheet_names():
            log.warn("Sheet " + activityLogSheetName + " will be created")
            work_sheet = work_book.create_sheet(activityLogSheetName)
        else:
            work_sheet = work_book.get_sheet_by_name(activityLogSheetName)
        self.__fill_sheet_with_default(work_sheet)
        return work_sheet

    
    def __get_first_activity_date(self, date_activities_map: WeeksActivities) -> datetime.date:
        first_workout_date = None
        for date in date_activities_map.keys():
            if not first_workout_date:
                first_workout_date = date
            elif date < first_workout_date:
                first_workout_date = date
        return first_workout_date if first_workout_date else datetime.date.today()

    def __fill_dates_column(self, work_sheet: Worksheet, first_workout_date: datetime.date):
        future_weeks = 12

        b3Value = work_sheet["B3"].value
        today = datetime.date.today()
        last_tracked_monday = today + datetime.timedelta(days=7 * future_weeks - today.weekday())
        if b3Value is not None:
            # Move existing rows down to add place for new weeks on top
            dateFromB3 = datetime.datetime.strptime(str(b3Value), "%Y-%m-%d").date()
            number_of_weeks = 0
            while dateFromB3 < last_tracked_monday:
                dateFromB3 = dateFromB3 + datetime.timedelta(days=7)
                number_of_weeks = number_of_weeks + 1
            if dateFromB3 != last_tracked_monday:
                raise ValueError('Error! B3 value is not Monday date')
            work_sheet.move_range("A3:Z2000", rows=number_of_weeks, cols=0)
        iteration_date = last_tracked_monday
        iterator = 3
        #fill empty dates in B column
        while first_workout_date < (iteration_date + datetime.timedelta(days=7)):
            if work_sheet["B" + str(iterator)].value is None:
                work_sheet["B" + str(iterator)] = str(iteration_date)
            iterator = iterator + 1
            iteration_date = iteration_date - datetime.timedelta(days=7)

    #Basal metabolic rate (BMR) is often used interchangeably with resting metabolic rate (RMR)
    def __compute_bmr(self):
        if self.__configs[self.SEX_KEY] == "M":
            self.__bmr = 88.362 + (13.397 * self.__configs[self.WEIGHT_KEY]) + (4.799 * self.__configs[self.HEIGHT_KEY]) - (5.677 * self.__configs[self.AGE_KEY])
        else:   
            self.__bmr = 447.593 + (9.247 * self.__configs[self.WEIGHT_KEY]) + (3.098 * self.__configs[self.HEIGHT_KEY]) - (4.330 * self.__configs[self.AGE_KEY])

    def __fill_week_report(self, week_activities: list[DayActivities]):
        @dataclass
        class ReportObject:
            moving_time: datetime.timedelta = datetime.timedelta(0)
            distance: float = 0

        activities_dict : dict[str, ReportObject] = {}
        for day_activities in week_activities:
            for activity_key in day_activities.keys():
                for activity in day_activities[activity_key]: # strava activity objects for a given date
                    if isinstance(activity.moving_time, datetime.timedelta):
                        if activity.type not in activities_dict:
                            activities_dict[str(activity.type)] = ReportObject(activity.moving_time, self.__to_float(activity.distance))
                        else:
                            activities_dict[str(activity.type)].moving_time += datetime.timedelta(0, activity.moving_time.seconds)
                            activities_dict[str(activity.type)].distance += self.__to_float(activity.distance)
        result = ""
        for activity_type, activity in activities_dict.items():    
            distance_in_kms = round(float(activity.distance) / 1000, 3)
            result += activity_type + ":\n"
            result += "    Total moving time: " + str(activity.moving_time) + "\n"
            result += "    Distance: " + str(distance_in_kms) + "kms\n"
            if activity_type == "Run":
                result += "    Average pace: " + self.__get_pace(activity.moving_time.seconds, distance_in_kms)
        return result

    def __get_pace(self, moving_time_in_seconds, distance_in_kms) -> str:
        pace_in_seconds = round(moving_time_in_seconds / distance_in_kms, 0)
        extra_zero = "0" if int(pace_in_seconds % 60) < 10 else ""
        return str(int(pace_in_seconds / 60)) + ":" + extra_zero + str(int(pace_in_seconds % 60)) + "min/km\n"
    

    def __get_km_per_hour(self, speed_in_meters_per_second) -> str:
        return str(round(float(speed_in_meters_per_second * 3.6), 2)) + "km/h\n"

    def __get_pace_from_speed(self, speed_in_meters_per_second) -> str:
        return str(round(16.66666666 / float(speed_in_meters_per_second), 2)) + "min/km\n"
    
    def __calculate_caloires_burnt(self, activity: Activity) -> float:
        calories = 0
        if activity.calories is not None:
            calories = activity.calories
        if activity.average_speed is not None and activity.moving_time is not None and activity.distance is not None and isinstance(activity.moving_time, datetime.timedelta):
            if activity.type == "Run":  
                calories = float(self.__bmr * self.__to_float(activity.average_speed) * 3.6 * activity.moving_time.seconds) / (24 * 3600)
            elif activity.type == "Swim":
                calories = self.__bmr * 8.5 * activity.moving_time.seconds / (24 * 3600) 
            elif activity.type == "Ride":
                calories = float(self.__bmr * self.__to_float(activity.average_speed) * 3.6 * activity.moving_time.seconds) / (24 * 3600 * 3) 
        return self.__to_float(calories)
    
    def __create_comment_for_activity(self, activity: Activity, calories: float) -> str:
        comment_value = "  Calories:" + str(round(calories, 0)) + "\n"
        if activity.moving_time is not None and activity.distance is not None and activity.type == "Run" and isinstance(activity.moving_time, datetime.timedelta):
            distance_in_kms = round(self.__to_float(activity.distance) / 1000, 3)
            if distance_in_kms > 0:
                comment_value += "  Average pace: " + self.__get_pace(activity.moving_time.seconds, distance_in_kms)
        if activity.average_speed is not None and activity.type != "Run":
            comment_value += "  Average speed: " + self.__get_km_per_hour(activity.average_speed)
        if activity.max_speed and self.__to_int(activity.max_speed) != 0:
            if activity.type != "Run":
                comment_value += "  Max speed: " + self.__get_km_per_hour(activity.max_speed)
        if activity.moving_time is not None:
            comment_value += "  Moving time: " + str(activity.moving_time) + "\n"
        if activity.distance is not None:
            comment_value += "  Distance: " + str(round(self.__to_float(activity.distance) / 1000, 3)) + "kms\n"
        if activity.average_heartrate is not None:
            comment_value += f"  Average heart rate: {str(activity.average_heartrate)} bpm\n"

        comment_value += "\n"
        return comment_value
    
    def __are_2_values_close_in_percents(self, value1: float, value2: float, percent: float) -> bool:
        return abs(value1 - value2) < value1 * percent / 100
    
    def __update_zone_values(self, zone_values: list[float], split):
        if split.average_heartrate and split.moving_time and isinstance(split.moving_time, datetime.timedelta):
            if split.average_heartrate < self.__heart_rate_zones[0]:
                zone_values[0] += split.moving_time.seconds
            elif split.average_heartrate < self.__heart_rate_zones[1]:
                zone_values[1] += split.moving_time.seconds
            elif split.average_heartrate < self.__heart_rate_zones[2]:
                zone_values[2] += split.moving_time.seconds
            elif split.average_heartrate < self.__heart_rate_zones[3]:
                zone_values[3] += split.moving_time.seconds
            else:
                zone_values[4] += split.moving_time.seconds

    def __fill_effort_levels(self, work_sheet: Worksheet, row: int, week_activities: list[DayActivities]):
        #other recovery time is a specific column that is not a part of strava activities but could be manually filled
        other_recovery_time = 0
        if work_sheet.cell(row, self.OTHER_RECOVERY_TIME_COLUMN).value:
            other_recovery_time = self.__to_int(work_sheet.cell(row, self.OTHER_RECOVERY_TIME_COLUMN).value)
        
        log.debug("Checking if effort recalculations needed for row {row}")
        if work_sheet.cell(row, self.ZONE_1_COLUMN_NUMBER).value is not None and \
            work_sheet.cell(row, self.ZONE_2_COLUMN_NUMBER).value is not None and \
            work_sheet.cell(row, self.ZONE_3_COLUMN_NUMBER).value is not None and \
            work_sheet.cell(row, self.ZONE_4_COLUMN_NUMBER).value is not None and \
            work_sheet.cell(row, self.ZONE_5_COLUMN_NUMBER).value is not None:
            #This row is already filled but might be not filled with all activities
            total_time_from_cells = self.__to_int(work_sheet.cell(row, self.ZONE_1_COLUMN_NUMBER).value) + \
                                    self.__to_int(work_sheet.cell(row, self.ZONE_2_COLUMN_NUMBER).value) + \
                                    self.__to_int(work_sheet.cell(row, self.ZONE_3_COLUMN_NUMBER).value) + \
                                    self.__to_int(work_sheet.cell(row, self.ZONE_4_COLUMN_NUMBER).value) + \
                                    self.__to_int(work_sheet.cell(row, self.ZONE_5_COLUMN_NUMBER).value)
            total_time_for_all_activities = other_recovery_time * 60
            
            for day_activities in week_activities:
                for activity_date in day_activities.keys():
                    for activity in day_activities[activity_date]:
                        #we want to count only activities with average heartrate
                        if activity.moving_time and activity.average_heartrate and isinstance(activity.moving_time, datetime.timedelta):
                            total_time_for_all_activities += activity.moving_time.seconds
            if self.__are_2_values_close_in_percents(total_time_for_all_activities / 60, total_time_from_cells, 2):
                log.debug(f"We don't need to fill effort levels, they are already fully filled for row {row}")
                return
        #In the following code we go through activities 1 by 1 and calculate time spent in each zone
        zone_values_in_seconds = [other_recovery_time * 60, 0.0, 0.0, 0.0, 0.0]
        request_to_strava_made = False
        for day_activities in week_activities:
            for activity_date in day_activities.keys():
                for activity in day_activities[activity_date]:
                    # Makes sense to retrieve full activity only if it has average heartrate
                    if activity.average_heartrate:
                        full_activity = self.__stravaService.retrieve_full_activity(activity.id)
                        request_to_strava_made = True
                        if full_activity.laps and isinstance(full_activity.laps, list) and len(full_activity.laps) > 3:
                            #There are many laps we can measure efforts by laps
                            log.info(f"Lap case {activity.name}")
                            for lap in full_activity.laps:
                                self.__update_zone_values(zone_values_in_seconds, lap)                                    
                        elif full_activity.splits_metric and isinstance(full_activity.splits_metric, list):
                            #There are more splits than laps, we can measure efforts by splits
                            log.info(f"Splits case {activity.name}")
                            for split in full_activity.splits_metric:
                                self.__update_zone_values(zone_values_in_seconds, split)
                                        
        work_sheet.cell(row, self.ZONE_1_COLUMN_NUMBER).value = round(zone_values_in_seconds[0] / 60, 1)
        work_sheet.cell(row, self.ZONE_2_COLUMN_NUMBER).value = round(zone_values_in_seconds[1] / 60, 1)
        work_sheet.cell(row, self.ZONE_3_COLUMN_NUMBER).value = round(zone_values_in_seconds[2] / 60, 1)
        work_sheet.cell(row, self.ZONE_4_COLUMN_NUMBER).value = round(zone_values_in_seconds[3] / 60, 1)
        work_sheet.cell(row, self.ZONE_5_COLUMN_NUMBER).value = round(zone_values_in_seconds[4] / 60, 1)
        if request_to_strava_made:
            log.warn(f"Effort levels filled for row {row}. You can add time sleep to avoid limit request on strava. max 1000/day 100/15min requests")
            self.__apply_style(work_sheet)
            self.__work_book.save(self.FILE_NAME)
        else:
            log.info(f"Effort levels filled for row {row}")


    #TODO: INVESTIGATE: IT IS CRUSHING EXCEL FOR SOME REASON
    def __set_cell_value_with_hyperlink(self, cell: Cell, activities: list[Activity], value: str):
        #Find the longest activity and put a link on it
        max_moving_time = 0
        longest_activity_id = 0
        for activity in activities:
            if activity.moving_time and isinstance(activity.moving_time, datetime.timedelta):
                max_moving_time = max(max_moving_time, activity.moving_time.seconds)
                longest_activity_id = activity.id
        if longest_activity_id != 0:
            hyperlink = "https://www.strava.com/activities/" + str(longest_activity_id)
            cell.value = f'=HYPERLINK("{hyperlink}", "{value}")'

    #returns a number which is approximate effort estimation in calories
    def __fill_day(self, cell: Cell, activities_for_day: list[Activity]):
        cell_value = ''
        comment_value = ''

        #Calculate callories and effort
        calories_burnt_in_day = 0
        for activity in activities_for_day: # list of strava activity objects for a given date
            #Cell content
            cell_value = f"{cell_value} {activity.type}: {activity.name}. Id: {activity.id}\n"

            #Calculate calories burnt
            calories = self.__calculate_caloires_burnt(activity)
            calories_burnt_in_day += calories

            #Add comment to the cell
            comment_value = comment_value + str(activity.type) + ": "
            comment_value += self.__create_comment_for_activity(activity, calories)
        

        if self.__should_fully_rewrite or cell.comment is None:
            cell.comment = Comment(comment_value, "ME")
            cell.comment.width = units.points_to_pixels(250)
            cell.comment.height = units.points_to_pixels(comment_value.count('\n') * 20)
        if self.__should_fully_rewrite or cell.value is None:
            # self.__set_cell_value_with_hyperlink(cell, activities_for_day, cell_value)
            cell.value = cell_value
        return calories_burnt_in_day

    def __fill_week(self, work_sheet: Worksheet, line_number: int, week_activities: list[DayActivities], monday_date: datetime.date):
        calories_str = ''
        for day_activities in week_activities:
            for activity_date in day_activities.keys():
                day_of_the_week = activity_date - monday_date
                cell = work_sheet.cell(line_number, 3 + day_of_the_week.days)
                activities_for_day = day_activities[activity_date]
                calories_burnt = 0
                if self.__should_fully_rewrite or cell.comment is None or cell.value is None or work_sheet.cell(line_number, self.CALORIES_COLUMN) is None:
                    #This is the case where some value is not present in the cell, so we need to fill it
                    calories_burnt = int(self.__fill_day(cell, activities_for_day))
                if calories_str == '':
                    calories_str = str(calories_burnt)
                else:
                    calories_str += "+" + str(calories_burnt) 
                #Looks like there is a bug in recent excel that resets size of comments
                if cell.comment:
                    cell.comment.width = units.points_to_pixels(250)
                    cell.comment.height = units.points_to_pixels(cell.comment.content.count('\n') * 20)
        week_report = self.__fill_week_report(week_activities)

        work_sheet.cell(line_number, self.REPORT_COLUMN).value = week_report
        if work_sheet.cell(line_number, self.CALORIES_COLUMN).value is None:
            work_sheet.cell(line_number, self.CALORIES_COLUMN).value = "=SUM(" + calories_str  + ")"
        
        #needs separate service calls to retrieve actions 1 by 1
        self.__fill_effort_levels(work_sheet, line_number, week_activities)

    def __fill_columns_with_values(self, work_sheet: Worksheet, date_activities_map : WeeksActivities):
        iterator = 3
        while work_sheet["B" + str(iterator)].value is not None:
            value_from_cell = work_sheet["B" + str(iterator)].value
            monday_date = datetime.datetime.strptime(str(value_from_cell), "%Y-%m-%d").date()
            if monday_date in date_activities_map:
                self.__fill_week(work_sheet, iterator, date_activities_map[monday_date], monday_date)
            iterator = iterator + 1

    def create_and_fill_tables(self, date_activities_map : WeeksActivities):
        self.__compute_bmr()
        work_sheet = self.__get_worksheet(self.__work_book)  
        self.__fill_dates_column(work_sheet, self.__get_first_activity_date(date_activities_map))
        self.__fill_columns_with_values(work_sheet, date_activities_map)
        self.__apply_style(work_sheet)
        self.__work_book.save(self.FILE_NAME)

    